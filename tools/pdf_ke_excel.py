"""
tools/pdf_ke_excel.py
==========================================
Ekstrak tabel dari PDF menjadi Excel untuk DIPERIKSA MANUAL.

    python tools/pdf_ke_excel.py berkas.pdf
    python tools/pdf_ke_excel.py berkas.pdf keluaran.xlsx

KENAPA KE EXCEL, BUKAN LANGSUNG KE JSON
---------------------------------------
Ini disengaja. Ekstraksi PDF menyisipkan kesalahan karakter yang tidak
terlihat sekilas. Contoh nyata dari berkas PDF mapping RSJPDHK: kode
luaran `L.01001` terbaca sebagai `L.0L1001`.

Kesalahan seperti itu berbahaya di data klinis karena diam. Pada master
3S, dua digit pertama kode luaran menentukan kategori DAN urutan
prioritas — kode yang rusak membuat diagnosis masuk kategori "Lainnya"
dan selalu diurutkan paling akhir, tanpa pesan error apa pun. Pada PPK,
kode yang rusak membuat entri tidak bisa ditemukan.

Karena itu alurnya sengaja dipaksa melewati pemeriksaan manusia:

    PDF  ->  Excel  ->  (PERIKSA & PERBAIKI)  ->  JSON  ->  validasi
             ^ tools/pdf_ke_excel.py              ^ tools/excel_ke_json.py

KALAU ANDA PUNYA BERKAS EXCEL ASLINYA, PAKAI ITU — jangan lewat PDF.
PDF adalah format cetak, bukan format data; sebagian informasi memang
hilang saat dokumen dicetak ke PDF dan tidak bisa dipulihkan.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

try:
    import pdfplumber
except ImportError:
    print("❌ Butuh pdfplumber. Install dengan: pip install pdfplumber")
    raise SystemExit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Butuh openpyxl. Install dengan: pip install openpyxl")
    raise SystemExit(1)


# =====================================================
# DETEKSI NILAI MENCURIGAKAN
# =====================================================
# Pola kode yang benar. Apa pun yang menyerupai kode tapi tidak cocok
# dengan pola ini ditandai untuk diperiksa.
_POLA_KODE_BENAR = re.compile(r"^(D\.\d{4}|L\.\d{5}|I\.\d{5}|PPK\.[A-Z]{2}\.\d{3}|LOKAL\.\d{3})$")
_MIRIP_KODE = re.compile(r"^[A-Z]{1,5}\.?[0-9A-Z.]{3,}$")


def _curiga(nilai: str) -> str | None:
    """Kembalikan alasan kalau nilai patut dicurigai, atau None."""
    teks = str(nilai or "").strip()
    if not teks:
        return None

    if _MIRIP_KODE.match(teks) and not _POLA_KODE_BENAR.match(teks):
        return "bentuk kode tidak lazim"

    # Huruf terselip di antara angka — pola khas kerusakan ekstraksi PDF,
    # persis seperti L.01001 -> L.0L1001.
    if re.search(r"\d[A-Za-z]\d", teks) and len(teks) < 20:
        return "huruf terselip di antara angka"

    # Spasi hilang antar-kata: huruf kecil langsung diikuti huruf besar.
    if re.search(r"[a-z]{3}[A-Z][a-z]{3}", teks):
        return "kemungkinan spasi hilang antar-kata"

    return None


def _rapikan(nilai) -> str:
    """
    Rapikan sel: satukan baris yang terpotong akibat pembungkusan kolom
    PDF, tapi PERTAHANKAN baris yang memang berupa butir terpisah.

    PDF memotong baris hanya karena lebar kolom, bukan karena pergantian
    butir — jadi baris yang diawali huruf kecil hampir pasti sambungan
    dari baris sebelumnya.
    """
    teks = str(nilai or "").strip()
    if not teks:
        return ""

    baris = [b.strip() for b in teks.split("\n")]
    hasil: list[str] = []
    for b in baris:
        if not b:
            continue
        if hasil and (b[0].islower() or hasil[-1].endswith(("-", ","))):
            hasil[-1] = f"{hasil[-1]} {b}".replace("- ", "")
        else:
            hasil.append(b)
    return "\n".join(hasil)


# =====================================================
# EKSTRAKSI
# =====================================================

def ekstrak(sumber: Path, tujuan: Path) -> int:
    wb = Workbook()
    wb.remove(wb.active)

    total_baris = 0
    total_curiga = 0
    ringkasan: list[tuple[int, int, int]] = []

    with pdfplumber.open(sumber) as pdf:
        print(f"Halaman: {len(pdf.pages)}\n")

        for nomor, halaman in enumerate(pdf.pages, start=1):
            tabel_list = halaman.extract_tables()
            if not tabel_list:
                print(f"  Halaman {nomor}: tidak ada struktur tabel — dilewati")
                continue

            for urut, tabel in enumerate(tabel_list, start=1):
                if not tabel or len(tabel) < 2:
                    continue

                judul = f"Hal{nomor}" + (f"_T{urut}" if len(tabel_list) > 1 else "")
                ws = wb.create_sheet(judul[:31])

                curiga_lembar = 0
                for baris in tabel:
                    bersih = [_rapikan(sel) for sel in baris]
                    ws.append(bersih)

                # Tandai sel yang patut diperiksa
                kuning = PatternFill("solid", fgColor="FFF2CC")
                merah = Font(color="C00000", bold=True)
                for baris_ws in ws.iter_rows(min_row=2):
                    for sel in baris_ws:
                        alasan = _curiga(sel.value)
                        if alasan:
                            sel.fill = kuning
                            sel.font = merah
                            sel.comment = None
                            curiga_lembar += 1

                # Gaya judul
                biru = PatternFill("solid", fgColor="2F5597")
                for sel in ws[1]:
                    sel.font = Font(bold=True, color="FFFFFF")
                    sel.fill = biru

                for idx in range(1, ws.max_column + 1):
                    ws.column_dimensions[get_column_letter(idx)].width = 45
                for baris_ws in ws.iter_rows():
                    for sel in baris_ws:
                        sel.alignment = Alignment(vertical="top", wrap_text=True)
                ws.freeze_panes = "A2"

                jumlah_data = len(tabel) - 1
                total_baris += jumlah_data
                total_curiga += curiga_lembar
                ringkasan.append((nomor, jumlah_data, curiga_lembar))
                print(f"  Halaman {nomor}: {jumlah_data} baris, "
                      f"{curiga_lembar} sel ditandai untuk diperiksa")

    if not wb.sheetnames:
        print("\n❌ Tidak ada tabel yang bisa diekstrak dari PDF ini.")
        print("   PDF ini kemungkinan hasil pindaian (gambar), bukan teks.")
        print("   Kalau begitu, ekstraksi otomatis tidak mungkin — masukkan data manual")
        print("   memakai template dari: python tools/json_ke_excel.py")
        return 0

    _tambah_catatan(wb, sumber, total_baris, total_curiga, ringkasan)

    tujuan.parent.mkdir(parents=True, exist_ok=True)
    wb.save(tujuan)

    print(f"\n✅ Hasil ekstraksi: {tujuan}")
    print(f"   {total_baris} baris, {total_curiga} sel ditandai kuning untuk diperiksa")
    return total_baris


def _tambah_catatan(wb, sumber: Path, total: int, curiga: int, ringkasan) -> None:
    ws = wb.create_sheet("BACA DULU", 0)
    isi = [
        ["⚠️  HASIL EKSTRAKSI PDF — WAJIB DIPERIKSA SEBELUM DIPAKAI"],
        [""],
        [f"Sumber      : {sumber.name}"],
        [f"Total baris : {total}"],
        [f"Sel ditandai: {curiga}"],
        [""],
        ["KENAPA HARUS DIPERIKSA"],
        [""],
        ["Ekstraksi PDF menyisipkan kesalahan karakter yang tidak terlihat sekilas."],
        ["Contoh nyata dari berkas mapping RSJPDHK: kode L.01001 terbaca L.0L1001."],
        [""],
        ["Kesalahan seperti itu tidak memunculkan pesan error. Pada master 3S, dua digit"],
        ["pertama kode luaran menentukan kategori dan urutan prioritas — kode yang rusak"],
        ["membuat diagnosis masuk kategori 'Lainnya' dan selalu diurutkan paling akhir."],
        [""],
        ["YANG DITANDAI KUNING"],
        [""],
        ["Sel berwarna kuning dengan tulisan merah = patut dicurigai:"],
        ["  - bentuk kode tidak lazim (mis. L.0L1001)"],
        ["  - huruf terselip di antara angka"],
        ["  - kemungkinan spasi hilang antar-kata"],
        [""],
        ["Penandaan ini hanya membantu — TIDAK menjamin semua kesalahan tertangkap."],
        ["Baca ulang isinya, terutama kode dan angka."],
        [""],
        ["LANGKAH BERIKUTNYA"],
        [""],
        ["1. Periksa dan perbaiki seluruh sel yang ditandai."],
        ["2. Susun ulang ke format kolom baku. Ambil templatnya dari:"],
        ["      python tools/json_ke_excel.py sdki   (atau ppk)"],
        ["   Salin isi yang sudah diperiksa ke kolom yang sesuai."],
        ["3. Impor: python tools/excel_ke_json.py <berkas.xlsx>"],
        ["4. Validasi: python tools/validasi_sdki.py   (atau validasi_ppk.py)"],
        [""],
        ["KALAU ANDA PUNYA BERKAS EXCEL ASLINYA, PAKAI ITU — jangan lewat PDF."],
        ["PDF adalah format cetak, bukan format data. Sebagian informasi memang hilang"],
        ["saat dokumen dicetak ke PDF dan tidak bisa dipulihkan."],
        [""],
        ["RINCIAN PER HALAMAN"],
        [""],
        ["Halaman", "Baris data", "Sel ditandai"],
    ]
    for baris in isi:
        ws.append(baris)
    for hal, jml, cur in ringkasan:
        ws.append([hal, jml, cur])

    ws.column_dimensions["A"].width = 92
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws["A1"].font = Font(bold=True, size=13, color="C00000")
    for baris_judul in (7, 16, 26, 40):
        if ws.cell(row=baris_judul, column=1).value:
            ws.cell(row=baris_judul, column=1).font = Font(bold=True)


def main() -> int:
    if len(sys.argv) < 2:
        print("Pemakaian: python tools/pdf_ke_excel.py <berkas.pdf> [keluaran.xlsx]")
        return 1

    sumber = Path(sys.argv[1])
    if not sumber.exists():
        print(f"❌ Berkas tidak ditemukan: {sumber}")
        return 1

    tujuan = Path(sys.argv[2]) if len(sys.argv) > 2 else sumber.with_suffix(".ekstrak.xlsx")

    print(f"Mengekstrak: {sumber}\n")
    jumlah = ekstrak(sumber, tujuan)
    if not jumlah:
        return 1

    print("\n⚠️  PERIKSA sel bertanda kuning sebelum dipakai — lihat lembar 'BACA DULU'.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
