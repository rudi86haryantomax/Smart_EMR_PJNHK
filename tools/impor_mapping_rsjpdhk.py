"""
tools/impor_mapping_rsjpdhk.py
==========================================
Impor berkas Excel `SDKI_SLKI_SIKI_mapping_RSJPDHK_*.xlsx` menjadi
master 3S JSON.

    python tools/impor_mapping_rsjpdhk.py <berkas.xlsx>
    python tools/impor_mapping_rsjpdhk.py <berkas.xlsx> keluaran.json

KENAPA ADA IMPORTER TERSENDIRI
------------------------------
`excel_ke_json.py` mengharapkan satu kolom per bagian (Kriteria Mayor,
Kriteria Minor, Faktor Risiko terpisah). Berkas kerja RSJPDHK memakai
tata letak yang berbeda dan lebih ringkas:

- **Satu kolom "Kriteria"** berisi ketiganya dengan penanda `mayor:`,
  `minor:`, dan `FR:`
- **Tidak ada kolom "Jenis"** — disimpulkan dari penanda tersebut
  (`FR:` berarti diagnosis Risiko)
- **Diagnosis tambahan tidak berkode**, dipisahkan oleh baris penanda
  "DX TAMBAHAN (TIDAK ADA DI SDKI)"
- Ada kolom **"Status Verifikasi"** yang tidak ada di skema JSON

Memaksa tim klinis menata ulang berkas kerjanya justru menghilangkan
manfaat konversi otomatis. Jadi importer ini menyesuaikan diri dengan
berkas aslinya, bukan sebaliknya.

Kolom "Status Verifikasi" ikut disimpan ke JSON. Isinya bermakna klinis
— entri bertanda "PERLU VERIFIKASI" adalah yang belum dipastikan
kesesuaiannya dengan SDKI resmi, dan itu perlu tetap terlihat setelah
konversi, bukan hilang di tengah jalan.
"""

from __future__ import annotations

import json
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ Butuh openpyxl. Install dengan: pip install openpyxl")
    raise SystemExit(1)

from _konversi import pecah_sel  # noqa: E402

# Judul kolom yang dicari (dicocokkan longgar, tahan beda spasi/baris baru).
_KOLOM = {
    "kode":         ["kode dx", "kode"],
    "nama":         ["nama diagnosis", "diagnosis"],
    "kriteria":     ["kriteria"],
    "luaran_kode":  ["kode luaran", "kode l"],
    "luaran_nama":  ["nama luaran"],
    "observasi":    ["intervensi (observasi)", "observasi"],
    "terapeutik":   ["intervensi (terapeutik)", "terapeutik"],
    "edukasi":      ["intervensi (edukasi)", "edukasi"],
    "kolaborasi":   ["intervensi (kolaborasi)", "kolaborasi"],
    "status":       ["status verifikasi", "status"],
    "catatan":      ["catatan"],
}

_PENANDA_TAMBAHAN = re.compile(r"dx\s*tambahan", re.IGNORECASE)


def _bersih(teks: Any) -> str:
    """Rapikan spasi dan baris baru dalam judul kolom."""
    return re.sub(r"\s+", " ", str(teks or "")).strip().lower()


def _peta_kolom(header: list[Any]) -> dict[str, int]:
    bersih = [_bersih(h) for h in header]
    peta: dict[str, int] = {}
    for kunci, kandidat in _KOLOM.items():
        for idx, judul in enumerate(bersih):
            if judul in kandidat:
                peta[kunci] = idx
                break
    return peta


# =====================================================
# PARSING KRITERIA
# =====================================================

def pecah_kriteria(teks: Any) -> dict[str, list[str]]:
    """
    Pecah sel "Kriteria" menjadi mayor / minor / faktor risiko.

    Format sumber:
        mayor:butir a, butir b, butir c
        minor:butir d, butir e
    atau untuk diagnosis risiko:
        FR:faktor a, faktor b

    Butir dipisah koma. Koma memang dipakai sebagai pemisah di berkas
    sumber, jadi frasa yang mengandung koma di dalamnya (mis. "mengi,
    wheezing") memang dimaksudkan sebagai dua butir terpisah.
    """
    hasil = {"mayor": [], "minor": [], "faktor_risiko": []}
    if not teks:
        return hasil

    isi = str(teks).strip()

    # Sisipkan pemisah baris sebelum tiap penanda, supaya penanda yang
    # ditulis menyambung pada baris yang sama tetap terdeteksi.
    isi = re.sub(r"\s*(mayor|minor|FR)\s*:", r"\n\1:", isi, flags=re.IGNORECASE)

    bagian_kini = None
    for baris in isi.split("\n"):
        baris = baris.strip()
        if not baris:
            continue

        cocok = re.match(r"^(mayor|minor|FR)\s*:\s*(.*)$", baris, re.IGNORECASE)
        if cocok:
            penanda = cocok.group(1).lower()
            bagian_kini = {"mayor": "mayor", "minor": "minor", "fr": "faktor_risiko"}[penanda]
            sisa = cocok.group(2).strip()
        else:
            sisa = baris

        if bagian_kini is None or not sisa:
            continue

        for butir in sisa.split(","):
            butir = butir.strip(" .;")
            if butir:
                # Huruf pertama dibesarkan agar seragam dengan gaya
                # penulisan butir lain di aplikasi.
                hasil[bagian_kini].append(butir[0].upper() + butir[1:])

    return hasil


def _kode_lokal(urutan: int) -> str:
    return f"LOKAL.{urutan:03d}"


# =====================================================
# IMPOR
# =====================================================

def impor(sumber: Path) -> tuple[list[dict[str, Any]], list[str]]:
    wb = load_workbook(sumber, data_only=True)
    ws = wb.worksheets[0]

    baris_semua = list(ws.iter_rows(values_only=True))
    if len(baris_semua) < 2:
        return [], ["Berkas kosong atau tidak punya baris data."]

    peta = _peta_kolom(list(baris_semua[0]))
    wajib = ["kode", "nama", "kriteria", "luaran_kode", "luaran_nama"]
    hilang = [k for k in wajib if k not in peta]
    if hilang:
        return [], [f"Kolom wajib tidak ditemukan: {hilang}. Judul terbaca: {baris_semua[0]}"]

    def ambil_sel(baris, kunci) -> Any:
        idx = peta.get(kunci)
        return baris[idx] if (idx is not None and idx < len(baris)) else None

    entri: list[dict[str, Any]] = []
    catatan_proses: list[str] = []
    mode_tambahan = False
    nomor_lokal = 0

    for nomor_baris, baris in enumerate(baris_semua[1:], start=2):
        nama = str(ambil_sel(baris, "nama") or "").strip()
        kode = str(ambil_sel(baris, "kode") or "").strip().upper()

        if not nama and not kode:
            continue

        # Baris penanda pemisah — bukan data.
        if _PENANDA_TAMBAHAN.search(nama):
            mode_tambahan = True
            catatan_proses.append(
                f"Baris {nomor_baris}: penanda 'DX TAMBAHAN' — "
                "baris berikutnya diperlakukan sebagai diagnosis non-SDKI."
            )
            continue

        if not nama:
            catatan_proses.append(f"Baris {nomor_baris}: tanpa nama diagnosis, dilewati.")
            continue

        is_sdki = bool(kode) and not mode_tambahan
        if not kode:
            nomor_lokal += 1
            kode = _kode_lokal(nomor_lokal)
            catatan_proses.append(
                f"Baris {nomor_baris}: '{nama}' tanpa kode -> diberi kode {kode}."
            )

        kriteria = pecah_kriteria(ambil_sel(baris, "kriteria"))
        jenis = "Risiko" if kriteria["faktor_risiko"] and not kriteria["mayor"] else "Aktual"

        item: dict[str, Any] = {
            "kode": kode,
            "nama": nama,
            "jenis": jenis,
            "is_sdki": is_sdki,
            "kriteria": kriteria,
            "luaran": {
                "kode": str(ambil_sel(baris, "luaran_kode") or "").strip(),
                "nama": str(ambil_sel(baris, "luaran_nama") or "").strip(),
            },
            "intervensi": {
                "observasi": pecah_sel(ambil_sel(baris, "observasi")),
                "terapeutik": pecah_sel(ambil_sel(baris, "terapeutik")),
                "edukasi": pecah_sel(ambil_sel(baris, "edukasi")),
                "kolaborasi": pecah_sel(ambil_sel(baris, "kolaborasi")),
            },
            "catatan": (str(ambil_sel(baris, "catatan") or "").strip() or None),
            "terkait": [],
        }

        status = str(ambil_sel(baris, "status") or "").strip()
        if status:
            item["status_verifikasi"] = status

        entri.append(item)

    return entri, catatan_proses


def periksa(entri: list[dict]) -> list[str]:
    masalah: list[str] = []
    terlihat: set[str] = set()

    for e in entri:
        kode = e["kode"]
        if kode in terlihat:
            masalah.append(f"{kode}: kode duplikat.")
        terlihat.add(kode)

        if not e["luaran"]["kode"]:
            masalah.append(f"{kode}: kode luaran kosong.")
        if not e["intervensi"]["observasi"]:
            masalah.append(f"{kode}: intervensi observasi kosong.")
        if e["jenis"] == "Aktual" and not e["kriteria"]["mayor"]:
            masalah.append(f"{kode}: jenis Aktual tapi kriteria mayor kosong.")
        if e["jenis"] == "Risiko" and not e["kriteria"]["faktor_risiko"]:
            masalah.append(f"{kode}: jenis Risiko tapi faktor risiko kosong.")

        prefix = e["luaran"]["kode"][:4]
        if prefix and not re.match(r"^L\.\d{2}$", prefix):
            masalah.append(
                f"{kode}: kode luaran '{e['luaran']['kode']}' bentuknya tidak lazim — "
                "periksa, karena prefix menentukan kategori dan prioritas."
            )

    return masalah


def main() -> int:
    if len(sys.argv) < 2:
        print("Pemakaian: python tools/impor_mapping_rsjpdhk.py <berkas.xlsx> [keluaran.json]")
        return 1

    sumber = Path(sys.argv[1])
    if not sumber.exists():
        print(f"❌ Berkas tidak ditemukan: {sumber}")
        return 1

    tujuan = Path(sys.argv[2]) if len(sys.argv) > 2 else ROOT / "data" / "sdki_slki_siki.json"

    print(f"Membaca: {sumber}\n")
    entri, catatan_proses = impor(sumber)

    if not entri:
        print("❌ Tidak ada data yang bisa diimpor.")
        for c in catatan_proses:
            print(f"   • {c}")
        return 1

    if catatan_proses:
        print("Catatan proses:")
        for c in catatan_proses:
            print(f"   • {c}")
        print()

    masalah = periksa(entri)
    if masalah:
        print(f"❌ {len(masalah)} masalah — impor DIBATALKAN, berkas JSON tidak diubah:\n")
        for m in masalah[:25]:
            print(f"   • {m}")
        return 1

    sdki = sum(1 for e in entri if e["is_sdki"])
    aktual = sum(1 for e in entri if e["jenis"] == "Aktual")
    perlu_verifikasi = [e["kode"] for e in entri
                        if "VERIFIKASI" in str(e.get("status_verifikasi", "")).upper()]

    print(f"✅ {len(entri)} diagnosis terbaca")
    print(f"   SDKI resmi / lokal : {sdki} / {len(entri) - sdki}")
    print(f"   Aktual / Risiko    : {aktual} / {len(entri) - aktual}")
    if perlu_verifikasi:
        print(f"   Perlu verifikasi   : {len(perlu_verifikasi)} → {', '.join(perlu_verifikasi)}")
    print()

    meta: dict[str, Any] = {}
    if tujuan.exists():
        try:
            meta = dict(json.loads(tujuan.read_text(encoding="utf-8")).get("meta", {}))
        except Exception:
            meta = {}
        cadangan = tujuan.with_suffix(f".backup-{datetime.now():%Y%m%d-%H%M%S}.json")
        shutil.copy2(tujuan, cadangan)
        print(f"📦 Cadangan berkas lama: {cadangan.name}")

    meta.pop("_sumber_file", None)
    meta["sumber"] = sumber.name
    meta["jumlah_diagnosis"] = len(entri)
    meta["diperbarui_pada"] = datetime.now().isoformat(timespec="seconds")
    meta.setdefault(
        "catatan_lisensi",
        "Redaksi kriteria dan intervensi merupakan adaptasi kerja internal RSJPDHK, "
        "bukan salinan verbatim buku SDKI/SLKI/SIKI PPNI. Verifikasi terhadap buku "
        "resmi PPNI tetap diperlukan sebelum dipakai sebagai acuan legal atau audit klinis.",
    )

    tujuan.parent.mkdir(parents=True, exist_ok=True)
    tujuan.write_text(
        json.dumps({"meta": meta, "diagnosis": entri}, ensure_ascii=False, indent=1),
        encoding="utf-8",
    )
    print(f"✅ Ditulis ke: {tujuan}\n")
    print("Langkah berikutnya:")
    print("   python tools/validasi_sdki.py")
    return 0


if __name__ == "__main__":
    sys.exit(main())
