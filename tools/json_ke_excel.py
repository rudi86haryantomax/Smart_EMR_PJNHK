"""
tools/json_ke_excel.py
==========================================
Ekspor master data JSON menjadi berkas Excel yang siap disunting.

    python tools/json_ke_excel.py                      # keduanya
    python tools/json_ke_excel.py sdki                 # master 3S saja
    python tools/json_ke_excel.py ppk                  # PPK saja
    python tools/json_ke_excel.py ppk keluaran.xlsx    # nama berkas sendiri

Alur kerja yang dimaksudkan:

    JSON  ->  Excel  ->  (sunting di Excel)  ->  JSON  ->  validasi
              ^ tools/json_ke_excel.py           ^ tools/excel_ke_json.py

Tiap butir kriteria/intervensi ditulis pada baris tersendiri DI DALAM
satu sel. Untuk menambah baris baru di dalam sel Excel, tekan Alt+Enter.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Butuh openpyxl. Install dengan: pip install openpyxl")
    raise SystemExit(1)

from _konversi import FORMAT, ambil, deteksi_format_json, gabung_sel  # noqa: E402

# Lebar kolom per jenis isi. Kolom daftar dibuat lebar karena isinya
# kalimat panjang; kolom kode dibuat sempit agar tabel tetap terbaca.
_LEBAR_DAFTAR = 52
_LEBAR_TEKS = 30
_LEBAR_KODE = 14


def _lebar(nama_kolom: str, adalah_daftar: bool) -> int:
    if adalah_daftar:
        return _LEBAR_DAFTAR
    if "kode" in nama_kolom.lower() or nama_kolom.lower() in {"jenis", "icd-10", "sdki resmi"}:
        return _LEBAR_KODE
    return _LEBAR_TEKS


def ekspor(format_: str, sumber: Path, tujuan: Path) -> int:
    doc = json.loads(sumber.read_text(encoding="utf-8"))

    terdeteksi = deteksi_format_json(doc)
    if terdeteksi and terdeteksi != format_:
        print(f"⚠️  Isi berkas terdeteksi sebagai '{terdeteksi}', bukan '{format_}'. "
              f"Memakai '{terdeteksi}'.")
        format_ = terdeteksi

    spek = FORMAT[format_]
    entri = doc.get(spek["kunci_daftar"], [])

    wb = Workbook()
    ws = wb.active
    ws.title = spek["sheet"]

    kolom = spek["kolom"]
    judul = [k[0] for k in kolom]
    ws.append(judul)

    gaya_judul = Font(bold=True, color="FFFFFF")
    latar = PatternFill("solid", fgColor="2F5597")
    for idx in range(1, len(judul) + 1):
        sel = ws.cell(row=1, column=idx)
        sel.font = gaya_judul
        sel.fill = latar
        sel.alignment = Alignment(vertical="center", wrap_text=True)

    for item in entri:
        baris = []
        for _, jalur, adalah_daftar in kolom:
            nilai = ambil(item, jalur)
            if adalah_daftar:
                baris.append(gabung_sel(nilai))
            elif isinstance(nilai, bool):
                baris.append("Ya" if nilai else "Tidak")
            else:
                baris.append("" if nilai is None else str(nilai))
        ws.append(baris)

    for idx, (nama_kolom, _, adalah_daftar) in enumerate(kolom, start=1):
        huruf = get_column_letter(idx)
        ws.column_dimensions[huruf].width = _lebar(nama_kolom, adalah_daftar)

    for baris in ws.iter_rows(min_row=2):
        for sel in baris:
            sel.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"  # judul tetap terlihat saat menggulir
    ws.auto_filter.ref = ws.dimensions

    # Lembar kedua berisi petunjuk, supaya orang yang menerima berkas ini
    # tanpa membaca README tetap tahu aturannya.
    _tambah_petunjuk(wb, format_, spek)

    tujuan.parent.mkdir(parents=True, exist_ok=True)
    wb.save(tujuan)
    print(f"✅ {len(entri)} entri diekspor ke: {tujuan}")
    return len(entri)


def _tambah_petunjuk(wb, format_: str, spek: dict) -> None:
    ws = wb.create_sheet("Petunjuk")
    baris = [
        ["PETUNJUK PENGISIAN"],
        [""],
        ["1. Satu baris = satu diagnosis. Jangan memecah satu diagnosis ke beberapa baris."],
        ["2. Untuk kolom berisi daftar (kriteria, intervensi, tatalaksana):"],
        ["   tulis tiap butir pada baris tersendiri DI DALAM satu sel."],
        ["   Tekan Alt+Enter untuk membuat baris baru di dalam sel."],
        ["3. Jangan mengubah judul kolom pada baris pertama — dipakai saat impor."],
        ["4. Kolom boleh dipindah urutannya; pencocokan memakai judul, bukan posisi."],
        ["5. Menambah diagnosis baru: tambahkan baris baru di bawah."],
        [""],
        ["SETELAH SELESAI MENYUNTING"],
        [""],
        [f"   python tools/excel_ke_json.py <berkas.xlsx>"],
        [f"   python tools/validasi_{format_}.py"],
        [""],
        ["Impor akan menolak dan melaporkan masalah bila ada kolom wajib yang kosong,"],
        ["kode duplikat, atau nilai yang tidak sesuai."],
        [""],
    ]

    if format_ == "sdki":
        baris += [
            ["ATURAN KHUSUS MASTER 3S"],
            [""],
            ["- Jenis harus 'Aktual' atau 'Risiko' (tulis persis)."],
            ["- Jenis Aktual WAJIB mengisi Kriteria Mayor."],
            ["- Jenis Risiko WAJIB mengisi Faktor Risiko."],
            ["- Kode Luaran menentukan kategori DAN urutan prioritas:"],
            ["    L.01 Respirasi · L.02 Sirkulasi · L.14 Keamanan · L.03 Nutrisi & Cairan"],
            ["    L.04 Eliminasi · L.08 Nyeri · L.05 Aktivitas · L.09 Ego · L.13 Penyuluhan"],
            ["  Salah memberi kode luaran membuat kategori dan prioritasnya ikut salah."],
            ["- Intervensi Observasi tidak boleh kosong."],
            ["- Kata pada kriteria menentukan usulan otomatis. Tulis juga istilah"],
            ["  sehari-hari yang lazim dipakai perawat agar tidak terlewat."],
            ["- Diagnosis di luar SDKI: pakai kode berawalan LOKAL. dan isi 'SDKI Resmi' = Tidak."],
        ]
    else:
        baris += [
            ["ATURAN KHUSUS PPK"],
            [""],
            ["- Anamnesis dan Tatalaksana Awal tidak boleh kosong."],
            ["- ICD-10 sebaiknya diisi untuk pencarian dan pelaporan."],
            ["- Urutan butir = urutan tampil. Taruh yang paling mendesak di atas."],
            ["- Bila menambah PPK kegawatan, daftarkan kodenya pada KODE_KRITIS"],
            ["  di services/ppk_service.py agar dapat penanda kritis."],
            ["- Dosis obat sebaiknya mengikuti formularium rumah sakit;"],
            ["  hindari menuliskan dosis yang belum diverifikasi."],
        ]

    for isi in baris:
        ws.append(isi)

    ws.column_dimensions["A"].width = 95
    ws["A1"].font = Font(bold=True, size=13)
    for sel in ("A11", "A18"):
        if ws[sel].value:
            ws[sel].font = Font(bold=True)


def main() -> int:
    argumen = sys.argv[1:]
    target = ["sdki", "ppk"]
    tujuan_khusus = None

    if argumen:
        pertama = argumen[0].strip().lower()
        if pertama in FORMAT:
            target = [pertama]
            if len(argumen) > 1:
                tujuan_khusus = Path(argumen[1])
        else:
            print(f"❌ Format tidak dikenal: '{argumen[0]}'. Pilih 'sdki' atau 'ppk'.")
            return 1

    total = 0
    for format_ in target:
        spek = FORMAT[format_]
        sumber = ROOT / "data" / spek["berkas_default"]
        if not sumber.exists():
            print(f"⚠️  Lewati '{format_}': {sumber} tidak ada.")
            continue
        tujuan = tujuan_khusus or (ROOT / "data" / f"{sumber.stem}.xlsx")
        total += ekspor(format_, sumber, tujuan)

    if total:
        print("\nSunting berkas Excel di atas, lalu impor kembali dengan:")
        print("   python tools/excel_ke_json.py <berkas.xlsx>")
    return 0


if __name__ == "__main__":
    sys.exit(main())
