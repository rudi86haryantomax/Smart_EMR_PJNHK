"""
tools/excel_ke_json.py
==========================================
Impor berkas Excel (atau CSV) menjadi master data JSON.

    python tools/excel_ke_json.py data/ppk_kardiovaskular.xlsx
    python tools/excel_ke_json.py berkas.xlsx keluaran.json
    python tools/excel_ke_json.py berkas.xlsx --format ppk

Format (sdki atau ppk) dideteksi otomatis dari judul kolom. Kolom
dicocokkan berdasarkan JUDUL, bukan posisi — jadi urutan kolom boleh
berbeda dan kolom tambahan buatan sendiri diabaikan tanpa merusak impor.

PENGAMAN: berkas JSON lama otomatis dicadangkan sebelum ditimpa, dan
impor dibatalkan bila ditemukan masalah yang membuat data tidak layak
pakai (kode kosong/duplikat, bagian wajib kosong). Lebih baik gagal
dengan pesan jelas daripada menimpa master data dengan isi yang rusak.
"""

from __future__ import annotations

import csv
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from _konversi import (  # noqa: E402
    FORMAT,
    deteksi_format_kolom,
    ke_bool,
    pecah_sel,
    taruh,
)


# =====================================================
# PEMBACAAN BERKAS
# =====================================================

def baca_excel(path: Path) -> tuple[list[str], list[list[Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("❌ Butuh openpyxl. Install dengan: pip install openpyxl")
        raise SystemExit(1)

    wb = load_workbook(path, data_only=True)

    # Lembar "Petunjuk" hasil ekspor bukan data — lewati.
    lembar = None
    for ws in wb.worksheets:
        if str(ws.title).strip().lower() not in {"petunjuk", "panduan", "instruksi"}:
            lembar = ws
            break
    if lembar is None:
        lembar = wb.worksheets[0]

    baris = list(lembar.iter_rows(values_only=True))
    if not baris:
        return [], []

    header = [str(h or "").strip() for h in baris[0]]
    isi = [list(b) for b in baris[1:]]
    return header, isi


def baca_csv(path: Path) -> tuple[list[str], list[list[Any]]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        pembaca = list(csv.reader(handle))
    if not pembaca:
        return [], []
    return [str(h or "").strip() for h in pembaca[0]], [list(b) for b in pembaca[1:]]


# =====================================================
# KONVERSI
# =====================================================

def _peta_kolom(header: list[str], kolom_spek: list[tuple[str, str, bool]]) -> dict[str, int]:
    """Petakan jalur JSON -> indeks kolom, berdasarkan kecocokan judul."""
    bersih = {str(h or "").strip().lower(): i for i, h in enumerate(header)}
    peta: dict[str, int] = {}
    for nama_kolom, jalur, _ in kolom_spek:
        idx = bersih.get(nama_kolom.strip().lower())
        if idx is not None:
            peta[jalur] = idx
    return peta


def _baris_kosong(baris: list[Any]) -> bool:
    return all(str(sel or "").strip() == "" for sel in baris)


def bangun_entri(
    baris: list[Any],
    kolom_spek: list[tuple[str, str, bool]],
    peta: dict[str, int],
    format_: str,
) -> dict[str, Any]:
    entri: dict[str, Any] = {}

    for _, jalur, adalah_daftar in kolom_spek:
        idx = peta.get(jalur)
        nilai = baris[idx] if (idx is not None and idx < len(baris)) else None

        if adalah_daftar:
            taruh(entri, jalur, pecah_sel(nilai))
        elif jalur == "is_sdki":
            taruh(entri, jalur, ke_bool(nilai))
        else:
            teks = str(nilai).strip() if nilai is not None else ""
            taruh(entri, jalur, teks or None)

    # Normalisasi bentuk agar selalu sama dengan JSON bawaan, meski
    # kolomnya tidak ada di berkas Excel.
    if format_ == "sdki":
        entri.setdefault("kriteria", {})
        for kunci in ("mayor", "minor", "faktor_risiko"):
            entri["kriteria"].setdefault(kunci, [])
        entri.setdefault("intervensi", {})
        for kunci in ("observasi", "terapeutik", "edukasi", "kolaborasi"):
            entri["intervensi"].setdefault(kunci, [])
        entri.setdefault("luaran", {})
        entri.setdefault("terkait", [])
        if "is_sdki" not in entri:
            entri["is_sdki"] = not str(entri.get("kode", "")).upper().startswith("LOKAL")
    else:
        entri.setdefault("kriteria", {})
        for kunci in ("anamnesis", "pemeriksaan_fisik", "penunjang", "kriteria_diagnosis"):
            entri["kriteria"].setdefault(kunci, [])
        entri.setdefault("tatalaksana", {})
        for kunci in ("awal", "farmakologis", "non_farmakologis", "rujukan"):
            entri["tatalaksana"].setdefault(kunci, [])
        entri.setdefault("edukasi", [])
        entri.setdefault("komplikasi", [])

    if entri.get("kode"):
        entri["kode"] = str(entri["kode"]).strip().upper()
    return entri


def periksa(entri_list: list[dict], format_: str) -> list[str]:
    """Pemeriksaan minimum sebelum menulis. Detail lengkap ada di validator."""
    masalah: list[str] = []
    terlihat: set[str] = set()

    for nomor, e in enumerate(entri_list, start=2):  # baris 1 = judul
        kode = e.get("kode")
        if not kode:
            masalah.append(f"Baris {nomor}: kolom Kode kosong.")
            continue
        if kode in terlihat:
            masalah.append(f"Baris {nomor}: kode '{kode}' duplikat.")
        terlihat.add(kode)

        if not e.get("nama"):
            masalah.append(f"Baris {nomor} ({kode}): Nama Diagnosis kosong.")

        if format_ == "sdki":
            jenis = str(e.get("jenis") or "").strip().lower()
            if jenis not in {"aktual", "risiko"}:
                masalah.append(
                    f"Baris {nomor} ({kode}): Jenis harus 'Aktual' atau 'Risiko' "
                    f"(isi: {e.get('jenis')!r})."
                )
            if jenis == "aktual" and not e["kriteria"]["mayor"]:
                masalah.append(f"Baris {nomor} ({kode}): Jenis Aktual wajib mengisi Kriteria Mayor.")
            if jenis == "risiko" and not e["kriteria"]["faktor_risiko"]:
                masalah.append(f"Baris {nomor} ({kode}): Jenis Risiko wajib mengisi Faktor Risiko.")
            if not e.get("luaran", {}).get("kode"):
                masalah.append(f"Baris {nomor} ({kode}): Kode Luaran kosong.")
            if not e["intervensi"]["observasi"]:
                masalah.append(f"Baris {nomor} ({kode}): Intervensi Observasi kosong.")
        else:
            if not e["kriteria"]["anamnesis"]:
                masalah.append(f"Baris {nomor} ({kode}): Anamnesis kosong.")
            if not e["tatalaksana"]["awal"]:
                masalah.append(f"Baris {nomor} ({kode}): Tatalaksana Awal kosong.")
            if not e.get("kategori"):
                masalah.append(f"Baris {nomor} ({kode}): Kategori kosong.")

    return masalah


def _meta_lama(tujuan: Path, format_: str, jumlah: int) -> dict[str, Any]:
    """
    Pertahankan meta dari JSON lama (termasuk peringatan lisensi dan
    peringatan draf), karena informasi itu tidak ada di Excel dan akan
    hilang kalau ditulis ulang dari nol.
    """
    meta: dict[str, Any] = {}
    if tujuan.exists():
        try:
            meta = dict(json.loads(tujuan.read_text(encoding="utf-8")).get("meta", {}))
        except Exception:
            meta = {}

    meta.pop("_sumber_file", None)
    meta["diperbarui_pada"] = datetime.now().isoformat(timespec="seconds")
    if format_ == "sdki":
        meta["jumlah_diagnosis"] = jumlah
    else:
        meta["jumlah"] = jumlah
    return meta


def main() -> int:
    argumen = [a for a in sys.argv[1:] if not a.startswith("--")]
    opsi = [a for a in sys.argv[1:] if a.startswith("--")]

    if not argumen:
        print("Pemakaian: python tools/excel_ke_json.py <berkas.xlsx|csv> [keluaran.json] [--format sdki|ppk]")
        return 1

    sumber = Path(argumen[0])
    if not sumber.exists():
        print(f"❌ Berkas tidak ditemukan: {sumber}")
        return 1

    format_paksa = None
    for o in opsi:
        if o.startswith("--format"):
            bagian = o.split("=", 1)
            format_paksa = bagian[1].strip().lower() if len(bagian) > 1 else None
    if format_paksa is None and "--format" in sys.argv:
        idx = sys.argv.index("--format")
        if idx + 1 < len(sys.argv):
            format_paksa = sys.argv[idx + 1].strip().lower()
            if format_paksa in argumen:
                argumen.remove(format_paksa)

    print(f"Membaca: {sumber}\n")

    if sumber.suffix.lower() in {".csv", ".tsv", ".txt"}:
        header, baris = baca_csv(sumber)
    else:
        header, baris = baca_excel(sumber)

    if not header:
        print("❌ Berkas kosong atau tidak punya baris judul.")
        return 1

    format_ = format_paksa or deteksi_format_kolom(header)
    if format_ not in FORMAT:
        print("❌ Format tidak dikenali dari judul kolom.")
        print(f"   Judul terbaca: {header[:6]}...")
        print("   Tentukan manual dengan: --format sdki   atau   --format ppk")
        return 1

    spek = FORMAT[format_]
    print(f"Format terdeteksi: {format_.upper()}")

    peta = _peta_kolom(header, spek["kolom"])
    dikenali = len(peta)
    total_kolom = len(spek["kolom"])
    print(f"Kolom dikenali   : {dikenali} dari {total_kolom}")

    hilang = [nama for nama, jalur, _ in spek["kolom"] if jalur not in peta]
    if hilang:
        print(f"⚠️  Kolom tidak ditemukan (akan dikosongkan): {', '.join(hilang)}")
    print()

    entri_list = [
        bangun_entri(b, spek["kolom"], peta, format_)
        for b in baris if not _baris_kosong(b)
    ]

    if not entri_list:
        print("❌ Tidak ada baris data.")
        return 1

    masalah = periksa(entri_list, format_)
    if masalah:
        print(f"❌ {len(masalah)} masalah — impor DIBATALKAN, berkas JSON tidak diubah:\n")
        for m in masalah[:25]:
            print(f"   • {m}")
        if len(masalah) > 25:
            print(f"   ... dan {len(masalah) - 25} lainnya")
        return 1

    tujuan = Path(argumen[1]) if len(argumen) > 1 else ROOT / "data" / spek["berkas_default"]

    if tujuan.exists():
        cadangan = tujuan.with_suffix(
            f".backup-{datetime.now():%Y%m%d-%H%M%S}.json"
        )
        shutil.copy2(tujuan, cadangan)
        print(f"📦 Cadangan berkas lama: {cadangan.name}")

    doc = {
        "meta": _meta_lama(tujuan, format_, len(entri_list)),
        spek["kunci_daftar"]: entri_list,
    }
    tujuan.parent.mkdir(parents=True, exist_ok=True)
    tujuan.write_text(json.dumps(doc, ensure_ascii=False, indent=1), encoding="utf-8")

    print(f"✅ {len(entri_list)} entri ditulis ke: {tujuan}\n")
    print("Langkah berikutnya — jalankan validasi lengkap:")
    print(f"   python tools/validasi_{format_}.py")
    return 0


if __name__ == "__main__":
    sys.exit(main())
